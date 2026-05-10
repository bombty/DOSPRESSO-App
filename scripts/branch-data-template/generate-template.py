"""
DOSPRESSO Şube Veri Toplama Excel Template Generator

KULLANIM:
  python scripts/branch-data-template/generate-template.py \
    --branch-id 5 \
    --branch-name "Antalya Işıklar" \
    --output /tmp/dospresso-sube-veri-isiklar.xlsx

ÇIKTI: 3 sheet'li Excel
  Sheet 1: 📖 Talimatlar (şube müdürü için açıklama)
  Sheet 2: 👥 Personel (boş, dropdown'larla)
  Sheet 3: 🔧 Ekipman (16 cihaz pre-listed, sadece doldur)

Aslan 10 May 2026 talebi.
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ═══════════════════════════════════════════════════════════════════
# SABITLER
# ═══════════════════════════════════════════════════════════════════

# DOSPRESSO renk paleti
COLOR_NAVY = "192838"      # Koyu lacivert (header)
COLOR_RED = "C0392B"       # Kırmızı (zorunlu uyarı)
COLOR_LIGHT_GRAY = "EDEAE4"  # Açık gri (background)
COLOR_WHITE = "FFFFFF"
COLOR_YELLOW = "FFF3CD"    # Sarı (dikkat)
COLOR_GREEN = "D4EDDA"     # Yeşil (tamam)

# Font
FONT_NAME = "Arial"

# 7 personel seviyesi (Aslan 10 May 2026)
PERSONNEL_LEVELS = [
    "Stajyer",
    "Bar Buddy",
    "Barista",
    "Supervisor Buddy",
    "Supervisor",
    "Müdür",
    "Partner",  # Yatırımcı/Owner — Aslan: "Partner"
]

# Çalışma tipi
EMPLOYMENT_TYPES = ["Full-time", "Part-time", "Stajyer", "Sezonluk"]

# 17 ekipman tipi (Aslan'ın listesi)
EQUIPMENT_TYPES = [
    {"key": "espresso", "name": "Espresso Makinesi (Thermoplan)", "default_count": None},
    {"key": "krema", "name": "Krema Makinesi", "default_count": None},
    {"key": "mixer", "name": "Bar Mikseri (Artemis)", "default_count": None},
    {"key": "blender", "name": "Blender (Blendtech)", "default_count": None},
    {"key": "cash", "name": "Kasa Sistemi", "default_count": None},
    {"key": "kiosk", "name": "Kiosk Sistemi", "default_count": None},
    {"key": "tea", "name": "Çay Makinesi", "default_count": None},
    {"key": "ice", "name": "Buz Makinesi (Manitowock)", "default_count": None},
    {"key": "filtre_kahve", "name": "Filtre Kahve Makinesi", "default_count": None},
    {"key": "turk_kahve", "name": "Türk Kahvesi Makinesi", "default_count": None},
    {"key": "donut_teshir", "name": "Donut Teşhir Dolabı", "default_count": None},
    {"key": "positive_teshir", "name": "+4 Teşhir Dolabı (Soğutmalı)", "default_count": None},
    {"key": "notr_teshir", "name": "Nötr Teşhir Dolabı (Oda Sıcaklığı)", "default_count": None},
    {"key": "teshir_set", "name": "Teşhir Dolabı Set", "default_count": None},
    {"key": "firin", "name": "Fırın", "default_count": None},
    {"key": "hizli_firin", "name": "Hızlı Fırın", "default_count": None},
    {"key": "tost", "name": "Tost Makinesi", "default_count": None},
]

# Bakım sıklığı (gün)
MAINTENANCE_INTERVALS = {
    "espresso": 30,
    "krema": 30,
    "mixer": 90,
    "blender": 60,
    "cash": 180,
    "kiosk": 90,
    "tea": 90,
    "ice": 60,
    "filtre_kahve": 90,
    "turk_kahve": 180,
    "donut_teshir": 180,
    "positive_teshir": 180,
    "notr_teshir": 365,
    "teshir_set": 180,
    "firin": 90,
    "hizli_firin": 90,
    "tost": 365,
}

# Teknik sorun seçenekleri (dropdown)
TECHNICAL_ISSUE_OPTIONS = [
    "Yok",
    "Hafif (kullanılabilir)",
    "Orta (acil servis)",
    "Ağır (çalışmıyor)",
]


# ═══════════════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ═══════════════════════════════════════════════════════════════════


def make_border(thickness: str = "thin") -> Border:
    """Hücre kenarlığı oluştur"""
    side = Side(border_style=thickness, color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def header_style() -> dict:
    """Başlık satırı stili"""
    return {
        "font": Font(name=FONT_NAME, bold=True, color=COLOR_WHITE, size=11),
        "fill": PatternFill("solid", start_color=COLOR_NAVY),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": make_border("medium"),
    }


def required_header_style() -> dict:
    """Zorunlu kolon başlığı (kırmızı *)"""
    return {
        "font": Font(name=FONT_NAME, bold=True, color=COLOR_WHITE, size=11),
        "fill": PatternFill("solid", start_color=COLOR_RED),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": make_border("medium"),
    }


def title_style() -> dict:
    """Sayfa başlık stili"""
    return {
        "font": Font(name=FONT_NAME, bold=True, color=COLOR_NAVY, size=16),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def info_style() -> dict:
    """Bilgi satırı stili"""
    return {
        "font": Font(name=FONT_NAME, size=10, italic=True),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }


def apply_style(cell, style: dict) -> None:
    """Stil objesini hücreye uygula"""
    if "font" in style:
        cell.font = style["font"]
    if "fill" in style:
        cell.fill = style["fill"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "border" in style:
        cell.border = style["border"]


# ═══════════════════════════════════════════════════════════════════
# SHEET 1 — TALIMATLAR
# ═══════════════════════════════════════════════════════════════════


def create_instructions_sheet(wb: Workbook, branch_name: str, branch_id: int) -> None:
    ws = wb.active
    ws.title = "📖 Talimatlar"

    # Başlık
    ws.merge_cells("A1:F1")
    ws["A1"] = "📋 DOSPRESSO ŞUBE VERİ TOPLAMA FORMU"
    apply_style(ws["A1"], title_style())
    ws.row_dimensions[1].height = 35

    # Şube bilgisi
    ws.merge_cells("A3:F3")
    ws["A3"] = f"Şube: {branch_name} (#{branch_id})  |  Tarih: {datetime.now().strftime('%d %B %Y')}"
    apply_style(ws["A3"], info_style())

    # Talimatlar
    instructions = [
        "",
        "🎯 BU FORMUN AMACI",
        "",
        "Pilot lansmanı öncesi (13 May 2026 Çarşamba 15:00) şubenizin tüm personel ve",
        "ekipman bilgilerini toplamak. Bu bilgiler:",
        "• Bordro hesaplamaları için (özlük, izin)",
        "• Bakım planlaması için (ekipman, garanti, son bakım)",
        "• Mr. Dobody bildirimleri için (geç gelme, bakım hatırlatma)",
        "• Yasal denetimler için (KVKK, İş Kanunu)",
        "kullanılır.",
        "",
        "📝 NASIL DOLDURULUR?",
        "",
        "1. SHEET 2 (👥 Personel) — Tüm aktif çalışanlarınızı sıralayın",
        "   • Her satıra 1 personel",
        "   • Kırmızı * olan kolonlar ZORUNLU",
        "   • Seviye + Tipi sütunlarında dropdown var (kendiniz yazmayın, seçin)",
        "",
        "2. SHEET 3 (🔧 Ekipman) — Şubenizdeki cihazlar listelenmiş",
        "   • 17 cihaz türü hazır listede var",
        "   • SADECE ŞUBENİZDE OLAN cihazları doldurun",
        "   • Olmayan cihazları boş bırakın (sistem 'yok' olarak yorumlar)",
        "   • 2 adet aynı cihaz varsa: 'Adet' kolonuna 2 yazın",
        "",
        "3. KAYDET ve YÜKLE",
        "   • Excel'i kaydet: dospresso-sube-veri-{şube-adı}.xlsx",
        "   • Sisteme yükle: Şube Müdürü Dashboard → 'Veri Yükle' butonu",
        "   • VEYA CGO'ya gönder: WhatsApp / email",
        "",
        "🚨 ÖNEMLİ NOTLAR",
        "",
        "• TC Kimlik No: 11 hane sayı (boş bırakmayın!)",
        "• Doğum Tarihi: GG.AA.YYYY formatında (örn: 15.05.1990)",
        "• IBAN: TR ile başlamalı (örn: TR12 3456 7890 1234 5678 9012 34)",
        "• Brüt Maaş: Sadece sayı (TL, ondalık virgül kullanın: 33.030,00)",
        "• Garanti / Bakım Tarihleri: GG.AA.YYYY formatında",
        "• Teknik Sorun: Cihaz çalışmıyorsa 'Ağır' seçin, hemen servise iletilir",
        "",
        "📞 SORUNUZ VARSA",
        "",
        "WhatsApp: CGO ile iletişime geçin",
        "Mr. Dobody Chat: Sistemden 'Yardım' butonuna tıklayın",
        "Aslan Bey: Bütün şubelerin bağlantı noktası",
        "",
        "✅ TAMAMLAYINCA",
        "",
        "Sistemde şu özellikler aktif olur:",
        "• Otomatik bordro hesaplama",
        "• Mr. Dobody bakım hatırlatıcı (14, 7, 1 gün önceden)",
        "• Garanti süresi takibi",
        "• Personel seviye yükseltme akışı",
        "• Otomatik raporlama (CGO için)",
        "",
        "💪 Bu form çok değerli — şubenizin etkin yönetimi için kritik!",
    ]

    for i, line in enumerate(instructions, start=5):
        ws.merge_cells(f"A{i}:F{i}")
        cell = ws[f"A{i}"]
        cell.value = line

        # Stil belirleme
        if line.startswith(("🎯", "📝", "🚨", "📞", "✅")):
            cell.font = Font(name=FONT_NAME, bold=True, size=12, color=COLOR_NAVY)
        elif line.startswith(("•", "1.", "2.", "3.")):
            cell.font = Font(name=FONT_NAME, size=10)
        elif line == "":
            pass
        else:
            cell.font = Font(name=FONT_NAME, size=10)

        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Kolon genişlikleri
    ws.column_dimensions["A"].width = 100


# ═══════════════════════════════════════════════════════════════════
# SHEET 2 — PERSONEL
# ═══════════════════════════════════════════════════════════════════


def create_personnel_sheet(wb: Workbook, branch_name: str, branch_id: int) -> None:
    ws = wb.create_sheet("👥 Personel")

    # Sayfa başlığı
    ws.merge_cells("A1:O1")
    ws["A1"] = f"PERSONEL BİLGİLERİ — {branch_name} (#{branch_id})"
    apply_style(ws["A1"], title_style())
    ws.row_dimensions[1].height = 30

    # Bilgi satırı
    ws.merge_cells("A2:O2")
    ws["A2"] = (
        "💡 Her satıra 1 personel. Kırmızı başlıklar ZORUNLU. "
        "Seviye + Tipi: dropdown'dan seçin. TC: 11 hane. IBAN: TR ile başlasın."
    )
    apply_style(ws["A2"], info_style())
    ws.row_dimensions[2].height = 25

    # Kolon başlıkları (3. satır)
    headers = [
        ("Ad *", True, 15),
        ("Soyad *", True, 15),
        ("TC No *", True, 15),
        ("Doğum Tarihi", False, 13),
        ("Giriş Tarihi *", True, 13),
        ("Tipi *", True, 12),
        ("Seviye *", True, 16),
        ("Yıllık İzin (gün)", False, 12),
        ("Brüt Maaş (TL)", False, 13),
        ("Banka", False, 15),
        ("IBAN", False, 28),
        ("Telefon", False, 14),
        ("Adres", False, 30),
        ("Tutanak / Uyarı", False, 25),
        ("Notlar", False, 25),
    ]

    for col_idx, (header, required, width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        if required:
            apply_style(cell, required_header_style())
        else:
            apply_style(cell, header_style())

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[3].height = 35

    # Boş satırlar (50 personel kapasiteli)
    for row_idx in range(4, 54):
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = make_border("thin")
            cell.font = Font(name=FONT_NAME, size=10)

    # Dropdown — Tipi (sütun F)
    dv_type = DataValidation(
        type="list",
        formula1=f'"{",".join(EMPLOYMENT_TYPES)}"',
        allow_blank=True,
    )
    dv_type.error = "Lütfen listeden seçin"
    dv_type.errorTitle = "Geçersiz Tip"
    ws.add_data_validation(dv_type)
    dv_type.add("F4:F53")

    # Dropdown — Seviye (sütun G)
    dv_level = DataValidation(
        type="list",
        formula1=f'"{",".join(PERSONNEL_LEVELS)}"',
        allow_blank=True,
    )
    dv_level.error = "Lütfen listeden seçin"
    dv_level.errorTitle = "Geçersiz Seviye"
    ws.add_data_validation(dv_level)
    dv_level.add("G4:G53")

    # Sayı kontrolü — TC (11 hane)
    dv_tc = DataValidation(
        type="textLength",
        operator="equal",
        formula1=11,
        allow_blank=True,
    )
    dv_tc.error = "TC Kimlik No 11 hane olmalı"
    dv_tc.errorTitle = "Geçersiz TC"
    ws.add_data_validation(dv_tc)
    dv_tc.add("C4:C53")

    # Tarih formatı — Doğum + Giriş
    for col_letter in ["D", "E"]:
        for row in range(4, 54):
            ws[f"{col_letter}{row}"].number_format = "DD.MM.YYYY"

    # Sayı formatı — Brüt Maaş, İzin
    for row in range(4, 54):
        ws[f"H{row}"].number_format = "0"
        ws[f"I{row}"].number_format = "#,##0.00"

    # Donmuş başlık
    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════
# SHEET 3 — EKIPMAN (PRE-LISTED 17 cihaz)
# ═══════════════════════════════════════════════════════════════════


def create_equipment_sheet(wb: Workbook, branch_name: str, branch_id: int) -> None:
    ws = wb.create_sheet("🔧 Ekipman")

    # Sayfa başlığı
    ws.merge_cells("A1:K1")
    ws["A1"] = f"EKİPMAN BİLGİLERİ — {branch_name} (#{branch_id})"
    apply_style(ws["A1"], title_style())
    ws.row_dimensions[1].height = 30

    # Bilgi satırı
    ws.merge_cells("A2:K2")
    ws["A2"] = (
        "💡 17 cihaz tipi listede hazır. "
        "Sadece ŞUBENİZDE OLAN cihazları doldurun. "
        "Olmayan cihazları boş bırakın. 2 adet varsa 'Adet' kolonuna 2 yazın."
    )
    apply_style(ws["A2"], info_style())
    ws.row_dimensions[2].height = 30

    # Kolon başlıkları
    headers = [
        ("Cihaz Adı (Pre-filled)", False, 35),
        ("Adet", False, 8),
        ("Marka / Model", False, 25),
        ("Model No", False, 18),
        ("Seri No", False, 20),
        ("Satınalma Tarihi", False, 15),
        ("Garanti Bitiş", False, 15),
        ("Son Bakım", False, 15),
        ("Sonraki Bakım (otomatik)", False, 18),
        ("Teknik Sorun", False, 18),
        ("Notlar", False, 30),
    ]

    for col_idx, (header, required, width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        apply_style(cell, header_style())
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[3].height = 35

    # 17 cihaz pre-listed (4-20. satır)
    for i, eq in enumerate(EQUIPMENT_TYPES, start=4):
        ws.cell(row=i, column=1, value=eq["name"])
        ws.cell(row=i, column=2, value=eq["default_count"])

        # Kategori bilgisi notlar'a
        # Bakım sıklığı (otomatik hesap için referans)
        ws.cell(
            row=i,
            column=11,
            value=f"Bakım sıklığı: {MAINTENANCE_INTERVALS[eq['key']]} gün",
        ).font = Font(name=FONT_NAME, size=9, italic=True, color="888888")

        # Cihaz adı kolonu — pre-filled, light gray
        cell_name = ws.cell(row=i, column=1)
        cell_name.fill = PatternFill("solid", start_color=COLOR_LIGHT_GRAY)
        cell_name.font = Font(name=FONT_NAME, size=10, bold=True)

        # Border
        for col_idx in range(1, 12):
            ws.cell(row=i, column=col_idx).border = make_border("thin")
            ws.cell(row=i, column=col_idx).font = Font(name=FONT_NAME, size=10)

        # Sonraki Bakım = Son Bakım + Bakım Sıklığı (formül)
        # Eğer son bakım dolduysa, otomatik hesaplanır
        ws.cell(
            row=i,
            column=9,
            value=f"=IF(H{i}=\"\",\"\",H{i}+{MAINTENANCE_INTERVALS[eq['key']]})",
        )

    # Boş satırlar (ekstra cihazlar için, 21-30)
    for row_idx in range(21, 31):
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = make_border("thin")
            cell.font = Font(name=FONT_NAME, size=10)

    # Dropdown — Teknik Sorun (sütun J)
    dv_issue = DataValidation(
        type="list",
        formula1=f'"{",".join(TECHNICAL_ISSUE_OPTIONS)}"',
        allow_blank=True,
    )
    dv_issue.error = "Lütfen listeden seçin"
    ws.add_data_validation(dv_issue)
    dv_issue.add("J4:J30")

    # Tarih formatı — Satınalma, Garanti, Son Bakım, Sonraki Bakım
    for col_letter in ["F", "G", "H", "I"]:
        for row in range(4, 31):
            ws[f"{col_letter}{row}"].number_format = "DD.MM.YYYY"

    # Sayı formatı — Adet
    for row in range(4, 31):
        ws[f"B{row}"].number_format = "0"

    # Conditional formatting — Teknik Sorun "Ağır" → kırmızı arka plan
    red_fill = PatternFill("solid", start_color="FFB3B3")
    ws.conditional_formatting.add(
        "J4:J30",
        CellIsRule(operator="equal", formula=['"Ağır (çalışmıyor)"'], fill=red_fill),
    )

    # Conditional formatting — Teknik Sorun "Orta" → sarı
    yellow_fill = PatternFill("solid", start_color="FFEB9C")
    ws.conditional_formatting.add(
        "J4:J30",
        CellIsRule(operator="equal", formula=['"Orta (acil servis)"'], fill=yellow_fill),
    )

    # Donmuş başlık
    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════
# ANA FONKSİYON
# ═══════════════════════════════════════════════════════════════════


def generate_template(branch_id: int, branch_name: str, output_path: str) -> None:
    """Excel template oluştur ve kaydet"""
    wb = Workbook()

    # 3 sheet oluştur (sırayla)
    create_instructions_sheet(wb, branch_name, branch_id)
    create_personnel_sheet(wb, branch_name, branch_id)
    create_equipment_sheet(wb, branch_name, branch_id)

    # İlk sheet'i aktif yap (talimatlar açılışta görünsün)
    wb.active = wb["📖 Talimatlar"]

    # Kaydet
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel template oluşturuldu: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="DOSPRESSO Şube Veri Excel Template")
    parser.add_argument("--branch-id", type=int, required=True, help="Şube ID")
    parser.add_argument("--branch-name", type=str, required=True, help="Şube adı")
    parser.add_argument(
        "--output",
        type=str,
        required=True,
        help="Çıktı dosya yolu (örn: /tmp/sube-veri.xlsx)",
    )
    args = parser.parse_args()

    try:
        generate_template(args.branch_id, args.branch_name, args.output)
    except Exception as e:
        print(f"❌ Hata: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
