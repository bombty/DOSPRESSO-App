"""
DOSPRESSO Şube Veri Excel Parse Script

KULLANIM:
  python scripts/branch-data-template/parse-template.py \
    --branch-id 8 \
    --input /tmp/upload-8-1234.xlsx

ÇIKTI: JSON (stdout)
  {
    "personnel": [
      {"firstName": "...", "lastName": "...", "tcNo": "...", ...},
      ...
    ],
    "equipment": [
      {"equipmentType": "espresso", "modelNo": "...", ...},
      ...
    ],
    "warnings": [
      {"sheet": "Personel", "row": 5, "field": "tcNo", "error": "..."}
    ]
  }
"""

import argparse
import json
import re
import sys
from datetime import datetime, date

from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════════
# Helper: Date parsing (Türkçe format DD.MM.YYYY veya Excel datetime)
# ═══════════════════════════════════════════════════════════════════


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # GG.AA.YYYY veya YYYY-MM-DD
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def normalize_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_int(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(float(s.replace(",", ".")))
        except ValueError:
            return None
    return None


def parse_decimal(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(".", "").replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


# ═══════════════════════════════════════════════════════════════════
# Sheet 2 — Personel parse
# ═══════════════════════════════════════════════════════════════════

# Excel kolon mapping (1-indexed)
PERSONNEL_COLUMNS = {
    "firstName": 1,    # Ad
    "lastName": 2,     # Soyad
    "tcNo": 3,         # TC No
    "birthDate": 4,    # Doğum Tarihi
    "hireDate": 5,     # Giriş Tarihi
    "employmentType": 6,  # Tipi
    "level": 7,        # Seviye
    "annualLeave": 8,  # Yıllık İzin
    "grossSalary": 9,  # Brüt Maaş
    "bank": 10,        # Banka
    "iban": 11,        # IBAN
    "phone": 12,       # Telefon
    "address": 13,     # Adres
    "warnings": 14,    # Tutanak
    "notes": 15,       # Notlar
}


def parse_personnel(ws):
    personnel = []
    warnings = []

    # Veri 4. satırdan başlar (1=başlık, 2=info, 3=kolon başlık)
    for row_idx in range(4, ws.max_row + 1):
        row = [
            ws.cell(row=row_idx, column=col).value
            for col in range(1, 16)
        ]

        # Tüm satır boşsa atla
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        person = {
            "firstName": normalize_str(row[0]),
            "lastName": normalize_str(row[1]),
            "tcNo": normalize_str(row[2]),
            "birthDate": parse_date(row[3]),
            "hireDate": parse_date(row[4]),
            "employmentType": normalize_str(row[5]),
            "level": normalize_str(row[6]),
            "annualLeave": parse_int(row[7]),
            "grossSalary": parse_decimal(row[8]),
            "bank": normalize_str(row[9]),
            "iban": normalize_str(row[10]),
            "phone": normalize_str(row[11]),
            "address": normalize_str(row[12]),
            "warnings": normalize_str(row[13]),
            "notes": normalize_str(row[14]),
        }

        # Validation
        if not person["firstName"] or not person["lastName"]:
            warnings.append({
                "sheet": "Personel",
                "row": row_idx,
                "field": "firstName/lastName",
                "error": "Ad veya Soyad eksik — atlandı",
            })
            continue

        # TC validation
        if person["tcNo"]:
            tc = person["tcNo"]
            if not re.match(r"^\d{11}$", tc):
                warnings.append({
                    "sheet": "Personel",
                    "row": row_idx,
                    "field": "tcNo",
                    "error": f"TC No 11 hane sayı olmalı: {tc}",
                })

        # IBAN validation
        if person["iban"]:
            iban = person["iban"].replace(" ", "")
            if not iban.startswith("TR") or len(iban) != 26:
                warnings.append({
                    "sheet": "Personel",
                    "row": row_idx,
                    "field": "iban",
                    "error": f"IBAN TR ile başlamalı + 26 hane: {iban}",
                })

        personnel.append(person)

    return personnel, warnings


# ═══════════════════════════════════════════════════════════════════
# Sheet 3 — Ekipman parse
# ═══════════════════════════════════════════════════════════════════

# Cihaz adı (Türkçe) → equipment_type (DB enum)
EQUIPMENT_NAME_TO_TYPE = {
    "Espresso Makinesi (Thermoplan)": "espresso",
    "Krema Makinesi": "krema",
    "Bar Mikseri (Artemis)": "mixer",
    "Blender (Blendtech)": "blender",
    "Kasa Sistemi": "cash",
    "Kiosk Sistemi": "kiosk",
    "Çay Makinesi": "tea",
    "Buz Makinesi (Manitowock)": "ice",
    "Filtre Kahve Makinesi": "filtre_kahve",
    "Türk Kahvesi Makinesi": "turk_kahve",
    "Donut Teşhir Dolabı": "donut_teshir",
    "+4 Teşhir Dolabı (Soğutmalı)": "positive_teshir",
    "Nötr Teşhir Dolabı (Oda Sıcaklığı)": "notr_teshir",
    "Teşhir Dolabı Set": "teshir_set",
    "Fırın": "firin",
    "Hızlı Fırın": "hizli_firin",
    "Tost Makinesi": "tost",
}


def parse_equipment(ws):
    equipment = []
    warnings = []

    # Veri 4. satırdan başlar
    for row_idx in range(4, ws.max_row + 1):
        row = [
            ws.cell(row=row_idx, column=col).value
            for col in range(1, 12)
        ]

        # Cihaz adı boşsa atla (boş satır)
        device_name = normalize_str(row[0])
        if not device_name:
            continue

        # Adet 0 veya boşsa atla (şube bu cihaza sahip değil)
        count = parse_int(row[1])
        if not count or count == 0:
            continue

        # Cihaz tipini map et
        equipment_type = EQUIPMENT_NAME_TO_TYPE.get(device_name)
        if not equipment_type:
            warnings.append({
                "sheet": "Ekipman",
                "row": row_idx,
                "field": "name",
                "error": f"Bilinmeyen cihaz adı: {device_name}",
            })
            continue

        item = {
            "equipmentType": equipment_type,
            "count": count,
            "brandModel": normalize_str(row[2]),
            "modelNo": normalize_str(row[3]),
            "serialNumber": normalize_str(row[4]),
            "purchaseDate": parse_date(row[5]),
            "warrantyEndDate": parse_date(row[6]),
            "lastMaintenanceDate": parse_date(row[7]),
            # Sonraki Bakım otomatik formül (Excel'de hesaplanmış olabilir)
            "nextMaintenanceDate": parse_date(row[8]),
            "technicalIssue": normalize_str(row[9]),
            "notes": normalize_str(row[10]),
        }

        equipment.append(item)

    return equipment, warnings


# ═══════════════════════════════════════════════════════════════════
# Ana fonksiyon
# ═══════════════════════════════════════════════════════════════════


def parse_template(branch_id: int, input_path: str) -> dict:
    wb = load_workbook(input_path, data_only=True)

    result = {
        "branchId": branch_id,
        "personnel": [],
        "equipment": [],
        "warnings": [],
    }

    # Sheet 2 — Personel
    if "👥 Personel" in wb.sheetnames:
        ws = wb["👥 Personel"]
        personnel, p_warnings = parse_personnel(ws)
        result["personnel"] = personnel
        result["warnings"].extend(p_warnings)
    else:
        result["warnings"].append({
            "sheet": "TÜMÜ",
            "row": 0,
            "field": "sheet",
            "error": "👥 Personel sheet'i bulunamadı",
        })

    # Sheet 3 — Ekipman
    if "🔧 Ekipman" in wb.sheetnames:
        ws = wb["🔧 Ekipman"]
        equipment, e_warnings = parse_equipment(ws)
        result["equipment"] = equipment
        result["warnings"].extend(e_warnings)
    else:
        result["warnings"].append({
            "sheet": "TÜMÜ",
            "row": 0,
            "field": "sheet",
            "error": "🔧 Ekipman sheet'i bulunamadı",
        })

    return result


def main():
    parser = argparse.ArgumentParser(description="DOSPRESSO Şube Veri Parse")
    parser.add_argument("--branch-id", type=int, required=True)
    parser.add_argument("--input", type=str, required=True)
    args = parser.parse_args()

    try:
        result = parse_template(args.branch_id, args.input)
        # JSON'u stdout'a yaz (Express child_process okuyacak)
        print(json.dumps(result, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
